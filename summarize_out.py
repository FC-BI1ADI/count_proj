# 导入统计汇总相关包
import numpy as np
import pandas as pd
# import matplotlib as mpl
# 导入地理编码模块
import compare_location as CL
# 导入OpenPyXL处理EXCEL的xlsx文件
import openpyxl

# add1(argument)
# 功能：通用加1操作，若值为None则设置值为1，否则值加1
# 参数：argument待加1值
# 返回：返回加1后的值
def add1(argument):
    if argument == None:
        return 1
    else:
        argument += 1
        return argument


# out_check(id,out_time,out_address)
# 功能：判断外出地址是否符合要求
# 参数：员工编号，外出时间，外出地址
# 返回：若同一天能查到2条同一位置的签卡记录则返回True，否则返回False
def out_check(oc_df, id, out_time, out_address):
    time_list = []
    for i in oc_df.index:
        oc_id = oc_df.loc[i, "员工编号"]
        oc_day = oc_df.loc[i, "签卡时间"][:10]
        oc_address = oc_df.loc[i, "地点"]
        if oc_id == id and oc_day == out_time and CL.compare_location(out_address, oc_address, 800) == 1:
            time_list.append(oc_df.loc[i, "签卡时间"][11:])
    # 判断查找到的记录数量，至少2个则结果为真，否则为假
    if len(time_list) >= 2:
        return True
    else:
        return False


# 调整显示格式
pd.set_option('display.max_columns', 10)
pd.set_option('display.max_rows', 1000)
pd.set_option('display.width', 200)

##################################################################################
# 处理外出记录单信息，形成df
##################################################################################
# 读入外出记录单文件
df_o = pd.read_excel("DATA/IN外出记录单.xlsx", header=2, usecols=[1, 2, 3, 4, 5, 6, 9, 11, 23])
# 调整各列的顺序
order_o = ["部门", "员工编号", "姓名", "人员类别", "外出时间", "外出类型", "外出地址", "相关项目编号", "拜访客户类型"]
df_o = df_o[order_o]
# 增加外出校核列
df_o["项目类型"] = None
df_o["外出校核"] = None

# 扫描项目编号列，判断是否为pipeline项目、非pipeline项目、无项目编号
for i in df_o.index:
    project_type = ""

    if pd.isnull(df_o.loc[i, "相关项目编号"]):
        project_type = "无项目编号"
    else:
        if df_o.loc[i, "相关项目编号"][0:1] == 'P':
            project_type = "Pipeline项目"
        else:
            project_type = "非Pipeline项目"
    df_o.loc[i, "项目类型"] = project_type

# 读入外勤签卡记录
df_check = pd.read_excel("DATA/IN外勤签卡记录.xlsx", header=0, usecols=[1, 3, 4])
# print(df_check)

# 较验外出合规性
for i in df_o.index:
    print("[%s-%d] 正在对%s的外出记录进行校核..." % (i + 1, df_o.shape[0], df_o.loc[i, "姓名"]))
    if out_check(df_check, df_o.loc[i, "员工编号"], df_o.loc[i, "外出时间"], df_o.loc[i, "外出地址"]) == True:
        df_o.loc[i, "外出校核"] = True
    else:
        df_o.loc[i, "外出校核"] = False
# df_o中是完全的外勤记录信息
# print(df_o)
# df_o.to_excel("data/OUT外勤汇总.xlsx",columns=["部门","员工编号","姓名","人员类别","外出类型","项目类型","外出校核"])

##################################################################################
# 处理出差单信息，df_b-->df_grp_b+df_info_b==>df_grp_b(new)
##################################################################################
# 读入出差单文件
df_b = pd.read_excel("DATA/IN出差单.xlsx", header=2, usecols=[1, 2, 3, 5, 15])
order_b = ["部门", "员工编号", "姓名", "人员类别", "实际出差天数"]
df_b = df_b[order_b]
# 将员工编号设为索引
df_b = df_b.set_index("员工编号")
# print(df_b)
# 将技术售前和技术售后统一为技术
df_b.loc[(df_b["人员类别"].str.find("技术售前")!=-1) | (df_b["人员类别"].str.find("技术售后")!=-1),"人员类别"] = "技术"
# 按员工编号对出差记录进行分组求次数和求和
grp_b = df_b.groupby(by=["员工编号"])
df_grp_b = pd.DataFrame(grp_b["实际出差天数"].agg(['count','sum']))

# 根据df_b补齐分组后的基本信息
df_info_b = pd.DataFrame(df_b,columns=["部门", "姓名", "人员类别"])
# 去除df_info_b中重复数据(为避免df_info_b.loc取值时出错，因返回值可能为Series或str)
df_info_b.drop_duplicates(inplace=True)

# 增加基础信息列
df_grp_b["部门"] = None
df_grp_b["姓名"] = None
df_grp_b["人员类别"] = None
# 遍历df_info_b查找基础信息并补齐
for i in df_grp_b.index:
    for j in df_info_b.index:
        if i==j:
            df_grp_b.loc[i,"部门"] = df_info_b.loc[j,"部门"]
            df_grp_b.loc[i,"姓名"] = df_info_b.loc[j,"姓名"]
            df_grp_b.loc[i,"人员类别"] = df_info_b.loc[j,"人员类别"]
            break


##################################################################################
# 读入模板，进行计数分析
##################################################################################
# 使用openpyxl读入模板，填数后，输出到excel文件中
wb = openpyxl.load_workbook(filename="data/区域销售考勤统计及分析模板.xlsx")
ws_sale = wb['销售类']
ws_tech = wb['技术类']
ws_comm = wb['综合类']


# 逐行扫描外出记录，分别判断并填入
# ------------------------------------------------------------------
for row in df_o.index:
    if df_o.loc[row, '外出校核'] == True:
        # 根据人员类型先分销售类和技术类
        # =======处理销售类=======
        if df_o.loc[row, '人员类别'] == '销售':
            for r in range(2, ws_sale.max_row + 1):
                if ws_sale.cell(r, 2).value == df_o.loc[row, '姓名']:
                    # 外出次数累加
                    ws_sale.cell(r, 9).value = add1(ws_sale.cell(r, 9).value)
                    # 项目类型
                    if df_o.loc[row, '项目类型'] == 'Pipeline项目':
                        ws_sale.cell(r, 10).value = add1(ws_sale.cell(r, 10).value)
                    if df_o.loc[row, '项目类型'] == '非Pipeline项目':
                        ws_sale.cell(r, 11).value = add1(ws_sale.cell(r, 11).value)
                    if df_o.loc[row, '项目类型'] == '无项目编号':
                        ws_sale.cell(r, 12).value = add1(ws_sale.cell(r, 12).value)
                    # 根据外出类型和项目类型，判断填表列
                    # -------商务非正式交流-------
                    if df_o.loc[row, '外出类型'] == '商务非正式交流':
                        ws_sale.cell(r, 13).value = add1(ws_sale.cell(r, 13).value)
                        if df_o.loc[row, '项目类型'] == 'Pipeline项目':
                            ws_sale.cell(r, 14).value = add1(ws_sale.cell(r, 14).value)
                        if df_o.loc[row, '项目类型'] == '非Pipeline项目':
                            ws_sale.cell(r, 15).value = add1(ws_sale.cell(r, 15).value)
                        if df_o.loc[row, '项目类型'] == '无项目编号':
                            ws_sale.cell(r, 16).value = add1(ws_sale.cell(r, 16).value)
                        if df_o.loc[row, '拜访客户类型'] == '最终用户':
                            ws_sale.cell(r, 21).value = add1(ws_sale.cell(r, 21).value)
                        if df_o.loc[row, '拜访客户类型'] == '合作伙伴':
                            ws_sale.cell(r, 22).value = add1(ws_sale.cell(r, 22).value)
                    # -------其他-------
                    if df_o.loc[row, '外出类型'] == '其他':
                        ws_sale.cell(r, 17).value = add1(ws_sale.cell(r, 17).value)
                        if df_o.loc[row, '项目类型'] == '无项目编号':
                            ws_sale.cell(r, 18).value = add1(ws_sale.cell(r, 18).value)
                    # -------拜访客户类型-------
                    if df_o.loc[row, '拜访客户类型'] == '最终用户':
                        ws_sale.cell(r, 19).value = add1(ws_sale.cell(r, 19).value)
                    if df_o.loc[row, '拜访客户类型'] == '合作伙伴':
                        ws_sale.cell(r, 20).value = add1(ws_sale.cell(r, 20).value)
        # =======处理技术类=======
        if df_o.loc[row, '人员类别'] == '技术售前' or df_o.loc[row, '人员类别'] == '技术售后':
            for r in range(2, ws_tech.max_row + 1):
                if ws_tech.cell(r, 2).value == df_o.loc[row, '姓名']:
                    # 外出次数累加
                    ws_tech.cell(r, 9).value = add1(ws_tech.cell(r, 9).value)
                    # 项目类型
                    if df_o.loc[row, '项目类型'] == 'Pipeline项目':
                        ws_tech.cell(r, 10).value = add1(ws_tech.cell(r, 10).value)
                    if df_o.loc[row, '项目类型'] == '非Pipeline项目':
                        ws_tech.cell(r, 11).value = add1(ws_tech.cell(r, 11).value)
                    if df_o.loc[row, '项目类型'] == '无项目编号':
                        ws_tech.cell(r, 12).value = add1(ws_tech.cell(r, 12).value)
                    # 根据外出类型和项目类型，判断填表列
                    # -------商务非正式交流-------
                    if df_o.loc[row, '外出类型'] == '商务非正式交流':
                        ws_tech.cell(r, 13).value = add1(ws_tech.cell(r, 13).value)
                        if df_o.loc[row, '项目类型'] == 'Pipeline项目':
                            ws_tech.cell(r, 14).value = add1(ws_tech.cell(r, 14).value)
                        if df_o.loc[row, '项目类型'] == '非Pipeline项目':
                            ws_tech.cell(r, 15).value = add1(ws_tech.cell(r, 15).value)
                        if df_o.loc[row, '项目类型'] == '无项目编号':
                            ws_tech.cell(r, 16).value = add1(ws_tech.cell(r, 16).value)
                    # -------客户交流-------
                    if df_o.loc[row, '外出类型'] == '客户交流':
                        ws_tech.cell(r, 17).value = add1(ws_tech.cell(r, 17).value)
                        if df_o.loc[row, '项目类型'] == 'Pipeline项目':
                            ws_tech.cell(r, 18).value = add1(ws_tech.cell(r, 18).value)
                        if df_o.loc[row, '项目类型'] == '非Pipeline项目':
                            ws_tech.cell(r, 19).value = add1(ws_tech.cell(r, 19).value)
                        if df_o.loc[row, '项目类型'] == '无项目编号':
                            ws_tech.cell(r, 20).value = add1(ws_tech.cell(r, 20).value)
                    # -------投标相关活动-------
                    if df_o.loc[row, '外出类型'] == '投标相关活动':
                        ws_tech.cell(r, 21).value = add1(ws_tech.cell(r, 21).value)
                        if df_o.loc[row, '项目类型'] == 'Pipeline项目':
                            ws_tech.cell(r, 22).value = add1(ws_tech.cell(r, 22).value)
                        if df_o.loc[row, '项目类型'] == '非Pipeline项目':
                            ws_tech.cell(r, 23).value = add1(ws_tech.cell(r, 23).value)
                        if df_o.loc[row, '项目类型'] == '无项目编号':
                            ws_tech.cell(r, 24).value = add1(ws_tech.cell(r, 24).value)
                    # -------售前客户培训和售后客户培训-------
                    if df_o.loc[row, '外出类型'] == '售前客户培训' or df_o.loc[row, '外出类型'] == '售前客户培训':
                        ws_tech.cell(r, 25).value = add1(ws_tech.cell(r, 25).value)
                        if df_o.loc[row, '项目类型'] == 'Pipeline项目':
                            ws_tech.cell(r, 26).value = add1(ws_tech.cell(r, 26).value)
                        if df_o.loc[row, '项目类型'] == '非Pipeline项目':
                            ws_tech.cell(r, 27).value = add1(ws_tech.cell(r, 27).value)
                        if df_o.loc[row, '项目类型'] == '无项目编号':
                            ws_tech.cell(r, 28).value = add1(ws_tech.cell(r, 28).value)
                    # -------安装实施次数/首次安装-------
                    if df_o.loc[row, '外出类型'] == '首次安装':
                        ws_tech.cell(r, 29).value = add1(ws_tech.cell(r, 29).value)
                    # -------故障排除次数/售后现场服务-------
                    if df_o.loc[row, '外出类型'] == '售后现场服务':
                        ws_tech.cell(r, 30).value = add1(ws_tech.cell(r, 30).value)
                    # -------巡检次数/巡检服务-------
                    if df_o.loc[row, '外出类型'] == '巡检服务':
                        ws_tech.cell(r, 31).value = add1(ws_tech.cell(r, 31).value)
                    # -------其他-------
                    if df_o.loc[row, '外出类型'] == '其他':
                        ws_tech.cell(r, 32).value = add1(ws_tech.cell(r, 32).value)
                        if df_o.loc[row, '项目类型'] == '无项目编号':
                            ws_tech.cell(r, 33).value = add1(ws_tech.cell(r, 33).value)
    # 外出校验未通过，外出异常次数进行累计
    else:
        if df_o.loc[row, '人员类别'] == '销售':
            for r in range(2, ws_sale.max_row + 1):
                if ws_sale.cell(r, 2).value == df_o.loc[row, '姓名']:
                    ws_sale.cell(r, 8).value = add1(ws_sale.cell(r, 8).value)
        if df_o.loc[row, '人员类别'] == '技术售前' or df_o.loc[row, '人员类别'] == '技术售后':
            for r in range(2, ws_tech.max_row + 1):
                if ws_tech.cell(r, 2).value == df_o.loc[row, '姓名']:
                    ws_tech.cell(r, 8).value = add1(ws_tech.cell(r, 8).value)

# 遍历出差数据输出到销售人员表中
# ------------------------------------------------------------------
for i in df_grp_b.index:
    # 仅处理销售人员数据
    if df_grp_b.loc[i,'人员类别'] == '销售':
        for r in range(2, ws_sale.max_row + 1):
            if ws_sale.cell(r, 2).value == df_grp_b.loc[i, '姓名']:
                ws_sale.cell(r, 23).value = df_grp_b.loc[i,'count']
                ws_sale.cell(r, 24).value = df_grp_b.loc[i, 'sum']


# 读取OUT考勤报表，计算考勤异常次数
# ------------------------------------------------------------------
wb_check = openpyxl.load_workbook(filename="data/OUT考勤报表.xlsx")
ws_check = wb_check.active

for row_index in range(2, ws_check.max_row + 1):
    # 获取姓名和考勤异常次数
    name = ws_check.cell(row_index, 3).value
    count = 0
    for col_index in range(4, ws_check.max_column + 1):
        cell_str = str(ws_check.cell(row_index, col_index).value)
        if cell_str.find('<') != -1:
            count += 1
    # 扫描ws_sale加入考勤异常次数
    for r in range(2, ws_sale.max_row + 1):
        if ws_sale.cell(r, 2).value == name and count != 0:
            ws_sale.cell(r, 6).value = count
    # 扫描ws_tech加入考勤异常次数
    for r in range(2, ws_tech.max_row + 1):
        if ws_tech.cell(r, 2).value == name and count != 0:
            ws_tech.cell(r, 6).value = count

##################################################################################
# 输出到分析报表文件中
##################################################################################
# 设置活动表单
wb.active = ws_sale
# 输出文件
output_file = "data/OUT分析报表.xlsx"
wb.save(filename=output_file)
