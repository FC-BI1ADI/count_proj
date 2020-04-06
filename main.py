# 导入常用操作模块
import common
# 导入日期时间处理模块
import datetime
import time
# 导入正则表达式模块
import re
# 导入OpenPyXL处理EXCEL的xlsx文件
import openpyxl
# 导入地理编码模块
import compare_location as CL
# 导入日期处理模块
import date_calculator as DC

#############################################################################
# 提醒进行文件改名操作
#############################################################################
prompt = '''
#############################################################################
    温馨提示：
    1.请将4个导出数据文件改名并另存为EXCEL新版本文件（.xlsx）
        IN考勤签卡数据.xlsx
        IN外出记录单.xlsx
        IN外勤签卡记录.xlsx
        IN请假单.xlsx
    2.准备完毕后，请按<任意键>继续处理。
    3.处理完毕会生成 OUT考勤报表.xlsx 文件
#############################################################################
'''

# input(prompt)


#############################################################################
# 处理考勤打卡记录
# 数据源文件：IN考勤签卡数据.xlsx
# 字段：1.部门，2.员工编号，3.姓名，4.签卡时间，5.数据来源，6.创建时间
#############################################################################
print("=======> 处理考勤打卡记录 =======>")
#############################################################################
# 构建考勤签卡列表OCR_list (official checking recode)
# 字段：department, id, name, check_time, source
#############################################################################
OCR_list = []

# 读入考勤签卡记录关键字段至OCR_list中
wb_OCR = openpyxl.load_workbook(filename="data/IN考勤签卡数据.xlsx")
ws_OCR = wb_OCR.active
row_index = 1
# 从第2行开始扫描记录
for per_row in ws_OCR.iter_rows():
    if row_index > 1:
        department = per_row[0].value
        id = per_row[1].value
        name = per_row[2].value
        # 若为光头数据，默认为考勤机机卡，补充数据
        if len(per_row[3].value) == 16:
            per_row[3].value += "(考勤机)"
        # 根据签卡类型，处理check_time
        if re.findall(r'[(](.*?)[)]', per_row[3].value)[0] == "定点签卡":
            check_time = time.strptime(per_row[3].value[:16], "%Y-%m-%d %H:%M")
        if re.findall(r'[(](.*?)[)]', per_row[3].value)[0] == "考勤机":
            check_time = time.strptime(per_row[3].value[:16], "%Y-%m-%d %H:%M")
        if re.findall(r'[(](.*?)[)]', per_row[3].value)[0] == "外勤签卡":
            check_time = time.strptime(per_row[3].value[:16], "%Y-%m-%d %H:%M")
        if re.findall(r'[(](.*?)[)]', per_row[3].value)[0] == "上班卡":
            check_time = time.strptime(per_row[3].value[:10] + " 08:59", "%Y-%m-%d %H:%M")
        if re.findall(r'[(](.*?)[)]', per_row[3].value)[0] == "下班卡":
            check_time = time.strptime(per_row[3].value[:10] + " 17:31", "%Y-%m-%d %H:%M")
        source = re.findall(r'[(](.*?)[)]', per_row[4].value)[0]
        # 将读入数据处理后加入OCR_list中
        OCR_list.append([department, id, name, check_time, source])

    row_index += 1

#############################################################################
# 生成考勤记录列表
#############################################################################
check_list = []

# check_list
# 关键字段:1.部门 2.ID 3.姓名 4.类型（考勤/外勤） 5.签卡(check_time)[列表]
for i in range(0, len(OCR_list)):
    department = OCR_list[i][0]
    id = OCR_list[i][1]
    name = OCR_list[i][2]
    date = "%4d-%02d-%02d" % (OCR_list[i][3].tm_year, OCR_list[i][3].tm_mon, OCR_list[i][3].tm_mday)
    # 若列表为空，直接添加记录
    if len(check_list) == 0:
        check_list.append([department, id, name, date, "考勤", []])
        continue
    # 先扫描一遍签卡记录看看有没有同一人、同一天
    found = False
    for j in range(0, len(check_list)):
        if check_list[j][1] == id and check_list[j][3] == date:
            found = True
    if found == False:
        check_list.append([department, id, name, date, "考勤", []])
# 此时，check_list中已是聚合后的列表

# 再扫描打卡记录，将时间写入聚合后的列表中
for i in range(0, len(OCR_list)):
    for j in range(0, len(check_list)):
        date = "%4d-%02d-%02d" % (OCR_list[i][3].tm_year, OCR_list[i][3].tm_mon, OCR_list[i][3].tm_mday)
        if check_list[j][1] == OCR_list[i][1] and check_list[j][3] == date:
            check_list[j][5].append("%02d:%02d" % (OCR_list[i][3].tm_hour, OCR_list[i][3].tm_min))

# 处理外出记录单 和 外勤打卡记录
print("=======> 处理外出记录和外勤签卡记录 =======>")
#############################################################################
# 依据外出记录单.xlsx比对IN外勤打卡记录.xlsx
# 比对项目包括：同一人、同一天、同一地点、2条签卡记录
# 若缺少任何一要素，即判断外勤记录失效，若均符合向check_list添加1条外勤记录
# department, id, name, date, "外勤", [到达时间,离开时间]
#############################################################################
# 读入外勤打卡记录
# 数据源文件：IN外勤打卡记录.xlsx
# 字段：1.所在部门，2.员工编号，3.员工姓名，4.地点，5.签卡时间，6.备注，7.审核人，8.审核时间，9.状态，10.签卡设备
# 关键字段：1.员工编号 2.签卡时间 3.地点
# 构建外勤列表MCR_list (mobile checking recode)
# 字段：department, id, name, check_time, location
MCR_list = []

# 读入外勤打卡记录关键字段至MCR_list中
wb_MCR = openpyxl.load_workbook(filename="data/IN外勤签卡记录.xlsx")
ws_MCR = wb_MCR.active
row_index = 1
for per_row in ws_MCR.iter_rows():
    if row_index > 1:
        check_time = time.strptime(per_row[4].value, "%Y-%m-%d %H:%M")
        MCR_list.append([per_row[0].value, per_row[1].value, per_row[2].value, check_time, per_row[3].value])
    row_index += 1

# 读入外出记录单到OW_list中
# 数据源文件：IN外出记录单.xlsx
# 字段：
# 1.编号，2.姓名，3.部门，4.外出时间，5.外出类型，(需加入员工编号）
# 6.人员类别，7.外出单位，8.外出地址9.相关项目名称，10.相关项目编号，
# 11.拜访人，12.相关销售人员，13.相关合同编号，14.联系人，15.联系方式，
# 16.外出具体目的，17.外出结果，18.审批，19.创建时间，20.流程状态
OW_list = []
wb_outwork = openpyxl.load_workbook(filename="data/IN外出记录单.xlsx")
ws_outwork = wb_outwork.active

# 读和外出记录形成OW_list
row_index = 1
for per_row in ws_outwork.iter_rows():
    if row_index > 3:
        OW_list.append(
            [per_row[3].value, str(int(per_row[2].value)), per_row[1].value, per_row[4].value, per_row[9].value, []])
    row_index += 1

# 比对外勤打卡记录MCR_list
for i in range(0, len(OW_list)):
    for j in range(0, len(MCR_list)):
        id = MCR_list[j][1]
        date = "%4d-%02d-%02d" % (MCR_list[j][3].tm_year, MCR_list[j][3].tm_mon, MCR_list[j][3].tm_mday)
        # 判断核心逻辑：如果同一人、同一天、同一地点，则添加记录到OW_list中
        if MCR_list[j][1] == OW_list[i][1] and date == OW_list[i][3] and CL.compare_location(MCR_list[j][4],
                                                                                             OW_list[i][4], 500) == 1:
            OW_list[i][5].append("%02d:%02d" % (MCR_list[j][3].tm_hour, MCR_list[j][3].tm_min))

# 添加OW_list中有效数据至check_list中
for i in range(0, len(OW_list)):
    # check_list : [department, id, name, date, "考勤", []
    # OW_list : [department, id, name, date, location, time[]]
    # 如果签卡次数 >= 2 ，则将所有签卡记录都加入列表中
    if len(OW_list[i][5]) > 1:
        check_list.append([OW_list[i][0], OW_list[i][1], OW_list[i][2], OW_list[i][3], "外勤", OW_list[i][5]])

#############################################################################
# 读入请假单，处理请假事项
#############################################################################
print("=======> 处理请假记录 =======>")
# 读入请假记录关键字段至AFL_list中（ask for leave)
AFL_list = []

wb_AFL = openpyxl.load_workbook(filename="data/IN请假单.xlsx")
ws_AFL = wb_AFL.active

row_index = 1
for per_row in ws_AFL.iter_rows():
    if row_index > 3:
        department = per_row[3].value
        # ID需要去除之前的0
        id = str(int(per_row[2].value))
        name = per_row[1].value
        type = per_row[6].value
        hst_time = time.strptime(per_row[7].value, "%Y-%m-%d %H:%M")
        het_time = time.strptime(per_row[8].value, "%Y-%m-%d %H:%M")
        AFL_list.append([department, id, name, type, hst_time, het_time])
    row_index += 1

# 分解每条假期记录，作为休假记录插入HD_list
# 字段：department, id, name, date, type, htime
# htime 为XX:XX-XX:XX格式的时间字符串
HD_list = []

for AFL_per_row in AFL_list:
    department = AFL_per_row[0]
    id = AFL_per_row[1]
    name = AFL_per_row[2]
    type = AFL_per_row[3]
    hst_time = AFL_per_row[4]
    het_time = AFL_per_row[5]
    days = DC.interval_day(hst_time, het_time) + 1

    # 根据days的不同，分别处理获取当前日期和每天覆盖时间段
    if days == 1:
        cur_day_str = "%04d-%02d-%02d" % (hst_time.tm_year, hst_time.tm_mon, hst_time.tm_mday)
        h_time = "%02d:%02d-%02d:%02d" % (hst_time.tm_hour, hst_time.tm_min, het_time.tm_hour, het_time.tm_min)
        HD_list.append([department, id, name, cur_day_str, type, h_time])
    if days == 2:
        # 2天中的首日
        cur_day_str = "%04d-%02d-%02d" % (hst_time.tm_year, hst_time.tm_mon, hst_time.tm_mday)
        h_time = "%02d:%02d-17:31" % (hst_time.tm_hour, hst_time.tm_min)
        HD_list.append([department, id, name, cur_day_str, type, h_time])
        # 2天中的末日
        cur_day_str = "%04d-%02d-%02d" % (het_time.tm_year, het_time.tm_mon, het_time.tm_mday)
        h_time = "08:59-%02d:%02d" % (het_time.tm_hour, het_time.tm_min)
        HD_list.append([department, id, name, cur_day_str, type, h_time])
    if days > 2:
        for day_index in range(0, days):
            if day_index == 0:
                cur_day_str = "%04d-%02d-%02d" % (hst_time.tm_year, hst_time.tm_mon, hst_time.tm_mday)
                h_time = "%02d:%02d-17:31" % (hst_time.tm_hour, hst_time.tm_min)
            if day_index > 0 and day_index < days:
                cur_day = DC.calculate_n_day(hst_time, day_index)
                cur_day_str = "%04d-%02d-%02d" % (cur_day.tm_year, cur_day.tm_mon, cur_day.tm_mday)
                h_time = "08:59-17:31"
            if day_index == days - 1:
                cur_day_str = "%04d-%02d-%02d" % (het_time.tm_year, het_time.tm_mon, het_time.tm_mday)
                h_time = "08:59-%02d:%02d" % (het_time.tm_hour, het_time.tm_min)
            # 将请假分解后的时间段写入HD_list
            HD_list.append([department, id, name, cur_day_str, type, h_time])

#############################################################################
# 功能：将HD_list中分解记录加入check_list
# HD_list字段：department, id, name, date, type, time（字符串）
# check_list字段：department, id, name, date, type, time（列表）
#############################################################################
for per_row in HD_list:
    time_list = []
    time_list.append(per_row[5][:5])
    time_list.append(per_row[5][6:])
    check_list.append([per_row[0], per_row[1], per_row[2], per_row[3], per_row[4], time_list])

# 聚合check_list形成 day_list
#############################################################################
# day_list字段如下：
# department, id, name, date, rec, status, reason
# rec为记录签卡情况的列表，status(正常|异常），reason为异常原因（迟到｜早退｜缺勤）
# check_list : department, id, name, date, type, time[]
day_list = []

for i in range(0, len(check_list)):

    department = check_list[i][0]
    id = check_list[i][1]
    name = check_list[i][2]
    date = check_list[i][3]

    # 如果day_list为空，则直接添加记录
    if len(day_list) < 1:
        day_list.append([department, id, name, date, [], "", ""])

    # 先扫描一遍day_list看看有没有同一人、同一天
    found = False
    for j in range(0, len(day_list)):
        if day_list[j][1] == id and day_list[j][3] == date:
            found = True
    if found == False:
        day_list.append([department, id, name, date, [], "", ""])
    # 至此已形成day_list的聚合表

    # 对check_list打卡时间列表进行预处理
    check_list[i][5].sort()
    n = len(check_list[i][5])
    # 根据check_time时间列表进行判断，若为1则缺少签卡纪录，判断属于上下班签卡，若>1,取最小值和最大值
    if check_list[i][4] == "考勤":
        if n == 1 and int(check_list[i][5][0][0:2]) < 12:
            rec_item = "考勤(%s-XX:XX)" % (check_list[i][5][0])
        if n == 1 and int(check_list[i][5][0][0:2]) >= 12:
            rec_item = "考勤（XX:XX-%s)" % (check_list[i][5][0])
        if n > 1:
            rec_item = "考勤(%s-%s)" % (check_list[i][5][0], check_list[i][5][n - 1])
    if check_list[i][4] == "外勤":
        rec_item = "外勤(%s-%s)" % (check_list[i][5][0], check_list[i][5][n - 1])
    if check_list[i][4] != "考勤" and check_list[i][4] != "外勤":
        rec_item = "%s(%s-%s)" % (check_list[i][4], check_list[i][5][0], check_list[i][5][n - 1])

    # 再次扫描day_list，如果同一人、同一天那么合并rec字段
    for j in range(0, len(day_list)):
        if day_list[j][1] == id and day_list[j][3] == date:
            day_list[j][4].append(rec_item)

# 至此，形成合并后的day_list

# 扫描获取考勤开始日期和结束日期
for i in range(0, len(day_list)):
    if i == 0:
        start_date = day_list[i][3]
        end_date = day_list[i][3]
    else:
        if day_list[i][3] < start_date:
            start_date = day_list[i][3]
        if day_list[i][3] > end_date:
            end_date = day_list[i][3]
# print(start_date, end_date)
# 根据开始和结束日期，生成日期范围列表date_list
date_list = common.get_dates_bytimes(start_date, end_date)

# 扫描获取user_list
# user_list : department, id, name
user_list = []
for i in range(0, len(day_list)):
    if len(user_list) == 0:
        user_list.append([day_list[i][0], day_list[i][1], day_list[i][2]])
    found = False
    for j in range(0, len(user_list)):
        if user_list[j][1] == day_list[i][1]:
            found = True
    if found == False:
        user_list.append([day_list[i][0], day_list[i][1], day_list[i][2]])

#############################################################################
# 判断day_list中数据是否存在异常，此部分判断逻辑是考核的关键
#############################################################################
for per_row in day_list:
    AM = "出勤"
    PM = "出勤"
    min_time = "12:00"
    max_time = "12:00"
    for per_item in per_row[4]:
        type = per_item[0:2]
        start_time = per_item[3:8]
        end_time = per_item[9:14]
        # 计算当日最小值和最大值
        if start_time < min_time:
            min_time = start_time
        if end_time > max_time and end_time != "XX:XX":
            max_time = end_time
        if type == "考勤":
            if start_time == "XX:XX":
                AM = "<缺勤>"
            if end_time == "XX:XX":
                PM = "<缺勤>"

    # 根据当日最小值和最大值来清除标志位
    if min_time == "12:00":
        AM = "<缺勤>"
    if max_time == "12:00":
        PM = "<缺勤>"
    if min_time < "09:05" and max_time >= "17:30":
        AM = "出勤"
        PM = "出勤"
    # 处理新疆地区的记录
    if per_row[0].find("新疆") != -1:
        if min_time > "10:30" and min_time != "12:00":
            AM = "<迟到>"
        if max_time < "19:00" and max_time != "12:00":
            PM = "<早退>"
    else:
        if min_time > "09:05" and min_time != "12:00":
            AM = "<迟到>"
        if max_time < "17:30" and max_time != "12:00":
            PM = "<早退>"
    # 单独处理外勤记录
    for per_item in per_row[4]:
        type = per_item[0:2]
        start_time = per_item[3:8]
        end_time = per_item[9:14]
        if type == "外勤":
            if start_time < "13:00":
                AM = "出勤"
            if end_time > "14:00":
                PM = "出勤"
    # 若上午或下午的标志位不是出勤，那么标记异常
    if AM != "出勤" or PM != "出勤":
        if AM != "出勤":
            per_row[5] = "异常"
            per_row[6] += AM
        if PM != "出勤":
            per_row[5] = "异常"
            per_row[6] += PM

    # print(per_row[3], AM, PM, start_time, end_time, min_time, max_time)
#############################################################################
# 输出已经过检验的day_list列表至EXCLE xlsx文件中
#############################################################################

# 创建考勤报表工作簿
wb_report = openpyxl.Workbook()
ws_report = wb_report.active
# 设置工作表格式


# 写入表头
header = ["部门", "员工编号", "姓名"] + date_list
ws_report.append(header)

# 写入用户信息
for per_row in user_list:
    ws_report.append(per_row)

# 读day_list列表，将信息写入单元格
# 定位单元格行row 列col

for per_row in day_list:
    for row_index in range(1, ws_report.max_row + 1):
        if per_row[1] == ws_report.cell(row_index, 2).value:
            break
    for col_index in range(4, ws_report.max_column + 1):
        if per_row[3] == ws_report.cell(1, col_index).value:
            break
    # print(row_index,col_index)

    # 拼接rec信息
    cell_str = ""
    for rec_item in per_row[4]:
        cell_str += rec_item
        cell_str += "\n"
    # 判断记录异常情况
    if per_row[5] == "异常":
        cell_str = per_row[6] + "\n" + cell_str
    # 写入单元格信息
    ws_report.cell(row_index, col_index).value = cell_str

#############################################################################
# 处理新疆考勤打卡问题（新疆地区员工上下班时间为：10:30-19:00
#############################################################################
# 扫描所有ws_report，若部门有“新疆”字样，依次扫描各列的考勤记录进行判断
# for row_index in range(2, ws_report.max_row + 1):
#     if ws_report.cell(row_index, 1).value.find("新疆") != -1:
#         for col_index in range(4, ws_report.max_column + 1):
#             cell_str = str(ws_report.cell(row_index, col_index).value)
#             # 取cell_str的考勤（XX:XX-XX:XX）字符串
#             if len(re.findall(r'考勤[(](.*?)[)]', cell_str)) >= 1:
#                 sub_str = re.findall(r'考勤[(](.*?)[)]', cell_str)[0]
#                 start_time,end_time = sub_str.split("-")
#                 # print(start_time,"$",end_time)
#                 if start_time < "10:30":
#                     cell_str.replace("<迟到>","")
#                 if end_time < "19:00":
#                     cell_str = "<早退>" + cell_str
#                 str(ws_report.cell(row_index, col_index).value = cell_str


# 标注单元格颜色
orange_fill = openpyxl.styles.PatternFill(fgColor="FFA500", fill_type='solid')
yellow_fill = openpyxl.styles.PatternFill(fgColor="FFFF00", fill_type='solid')
blue_fill = openpyxl.styles.PatternFill(fgColor="6495ED", fill_type='solid')
green_fill = openpyxl.styles.PatternFill(fgColor="9BCD9B", fill_type='solid')

# 扫描ws_report表，标注休息日
for col_index in range(4, ws_report.max_column + 1):
    date_header = time.strptime(ws_report.cell(1, col_index).value, "%Y-%m-%d")
    # tm_wday 取值0-6，0是周一，1是周二，2是周三，3是周四，4是周五，5是周六，6是周日
    # 如果是周末，那就标注为休息并将单元格标为蓝色
    if date_header.tm_wday == 5 or date_header.tm_wday == 6:
        for row_index in range(2, ws_report.max_row + 1):
            if ws_report.cell(row_index, col_index).value == None:
                ws_report.cell(row_index, col_index).value = "休息"
            else:
                ws_report.cell(row_index, col_index).value = "休息\n" + str(ws_report.cell(row_index, col_index).value)
            ws_report.cell(row_index, col_index).fill = blue_fill

# 扫描ws_report表，如果单元格内容为空，则意味着缺少考勤记录标记为橙色
for row_index in range(2, ws_report.max_row + 1):
    for col_index in range(4, ws_report.max_column + 1):
        cell_str = str(ws_report.cell(row_index, col_index).value)
        # print(row_index,col_index,"-",cell_str)
        if cell_str == "None":
            ws_report.cell(row_index, col_index).value = "缺勤"
            ws_report.cell(row_index, col_index).fill = yellow_fill
        if cell_str.find("<") != -1:
            ws_report.cell(row_index, col_index).fill = orange_fill
        # 自动换行设置
        ws_report.cell(row_index, col_index).alignment = openpyxl.styles.Alignment(wrapText=True)

# 设置列宽
ws_report.column_dimensions['A'].width = 20
ws_report.column_dimensions['B'].width = 10
ws_report.column_dimensions['C'].width = 10
for col_index in range(4, ws_report.max_column + 1):
    col_char = openpyxl.utils.get_column_letter(col_index)
    ws_report.column_dimensions[col_char].width = 20

# 向考勤报表工作簿中写入考勤信息
output_file = "data/OUT考勤报表.xlsx"
wb_report.save(filename=output_file)

print("======= 数据处理完毕=======")
